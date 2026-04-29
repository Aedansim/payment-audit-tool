"""
make_scoring_reference.py

Generates output/Scoring_Methodology.xlsx — a reference document explaining
the mathematics behind the Isolation Forest, Local Outlier Factor, and
Benford's Law scores that feed into the composite risk score.

Run from the project root:
    python make_scoring_reference.py
"""
from pathlib import Path
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------
NAVY    = "1F3864"
L_BLUE  = "D9E1F2"
F_BLUE  = "EBF1FA"   # formula cell background
ORANGE  = "ED7D31"
L_ORANGE= "FCE4D6"
GREEN   = "70AD47"
L_GREEN = "E2EFDA"
YELLOW  = "FFF2CC"
GREY    = "F2F2F2"
WHITE   = "FFFFFF"
ALT     = "F5F8FE"

THIN  = Side(style="thin",   color="C0C0C0")
MED   = Side(style="medium", color=NAVY)
TB    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_B = Border(left=MED,  right=MED,  top=MED,  bottom=MED)

# ---------------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------------

def _fill(c):
    return PatternFill("solid", fgColor=c)

def _font(size=10, bold=False, italic=False, color="000000", mono=False):
    return Font(size=size, bold=bold, italic=italic, color=color,
                name="Courier New" if mono else "Calibri")

def _al(h="left", v="top", wrap=True, ind=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=ind)

def _cell(ws, r, c, val, f=None, fl=None, al=None, b=None, span=1, h=None, nf=None):
    cell = ws.cell(row=r, column=c, value=val)
    if f:   cell.font      = f
    if fl:  cell.fill      = fl
    if al:  cell.alignment = al
    if b:   cell.border    = b
    if nf:  cell.number_format = nf
    if span > 1:
        ws.merge_cells(start_row=r, start_column=c,
                       end_row=r,   end_column=c + span - 1)
    if h:
        ws.row_dimensions[r].height = h
    return cell

# ---------------------------------------------------------------------------
# Composite helpers (each returns next row number)
# ---------------------------------------------------------------------------

SPAN = 8   # default merge span for full-width rows

def section(ws, r, text, color=NAVY):
    _cell(ws, r, 1, text,
          f=_font(12, bold=True, color=WHITE),
          fl=_fill(color), al=_al("left","center",wrap=False),
          span=SPAN, h=20)
    return r + 1

def subsection(ws, r, text, color=L_BLUE):
    _cell(ws, r, 1, text,
          f=_font(10, bold=True, color=NAVY),
          fl=_fill(color), al=_al("left","center",wrap=False),
          span=SPAN, h=16)
    return r + 1

def body(ws, r, text, bold=False, italic=False, color="000000", bg=None, h=None):
    _cell(ws, r, 1, text,
          f=_font(10, bold=bold, italic=italic, color=color),
          fl=_fill(bg) if bg else None,
          al=_al("left","top",wrap=True),
          span=SPAN)
    if h:
        ws.row_dimensions[r].height = h
    return r + 1

def formula(ws, r, text, h=16):
    _cell(ws, r, 1, text,
          f=_font(10, bold=True, color=NAVY, mono=True),
          fl=_fill(F_BLUE),
          al=_al("left","center",wrap=True,ind=1),
          span=SPAN, h=h)
    return r + 1

def note(ws, r, text):
    _cell(ws, r, 1, text,
          f=_font(9, italic=True, color="595959"),
          al=_al("left","top",wrap=True),
          span=SPAN)
    return r + 1

def blank(ws, r, h=6):
    ws.row_dimensions[r].height = h
    return r + 1

def hdr_row(ws, r, labels, fills=None, widths=None):
    for i, lbl in enumerate(labels, 1):
        bg = fills[i-1] if fills else NAVY
        _cell(ws, r, i, lbl,
              f=_font(9, bold=True, color=WHITE),
              fl=_fill(bg), al=_al("center","center",wrap=True),
              b=TB, h=28)
    return r + 1

def data_row(ws, r, values, bg=WHITE, bold_first=False, number_cols=None):
    for i, v in enumerate(values, 1):
        bf = (bold_first and i == 1)
        h_align = "left" if i == 1 else "center"
        _cell(ws, r, i, v,
              f=_font(9, bold=bf),
              fl=_fill(bg),
              al=_al(h_align, "center", wrap=True),
              b=TB)
    return r + 1

def widths(ws, w_list):
    for i, w in enumerate(w_list, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ===========================================================================
# Sheet 1 — Overview
# ===========================================================================

def build_overview(ws):
    widths(ws, [22, 10, 14, 20, 20, 20, 18, 18])

    r = 1
    _cell(ws, r, 1, "Payment Audit Tool — Scoring Methodology Reference",
          f=_font(14, bold=True, color=WHITE),
          fl=_fill(NAVY), al=_al("center","center",wrap=False),
          span=SPAN, h=30)
    r += 1
    r = body(ws, r,
             "This workbook documents the mathematical basis for the three analytical scores "
             "(Isolation Forest, Local Outlier Factor, Benford's Law) that, together with the "
             "Statistical Z-Score and Rule-Based Flags, combine into the composite risk score "
             "for each transaction line. It is intended as an audit-trail reference to support "
             "queries about how any individual risk score was derived.",
             italic=True, color="595959", h=40)
    r = blank(ws, r)

    # ---- Component summary table ----
    r = section(ws, r, "Component Scores — Summary")
    r = blank(ws, r, 4)
    r = hdr_row(ws, r, ["Scoring Component", "Weight", "Score Range",
                         "Anomaly Boundary", "Key Parameter(s)",
                         "Normalisation", "Sheet in This File", "Column in Pipeline"])
    rows_data = [
        ("Isolation Forest (IF)", "30 %", "[0, 1]",
         "predict() == −1  (top 5% of dataset)",
         "n_estimators=300, contamination=0.05",
         "Min-max across all transactions in run",
         "Sheet: Isolation Forest", "if_score"),
        ("Local Outlier Factor (LOF)", "25 %", "[0, 1]",
         "fit_predict() == −1  (top 5% of peer group)",
         "n_neighbors=20, contamination=0.05",
         "Min-max across all transactions in run",
         "Sheet: Local Outlier Factor", "lof_score"),
        ("Statistical Z-Score", "25 %", "[0, 1]",
         "max(|z_vendor|, |z_cc|) > 2.0  (2 std devs)",
         "Group mean & std per Vendor ID / Cost Centre",
         "Min-max across all transactions in run",
         "Sheet: Composite Score", "zscore_score"),
        ("Rule-Based Flags", "15 %", "[0, 1]",
         "Any of 8 binary rules triggered",
         "8 hard-coded forensic audit rules",
         "Fraction of 8 rules triggered (÷8)",
         "Sheet: Composite Score", "rule_flags_score"),
        ("Benford's Law", "5 %", "[0, 1]",
         "First digit among top-3 most deviant digits",
         "Non-recurring transactions only; suppression rule applies",
         "÷ max deviation in dataset",
         "Sheet: Benford's Law", "benford_score"),
    ]
    for i, row_d in enumerate(rows_data):
        bg = ALT if i % 2 == 0 else WHITE
        r = data_row(ws, r, row_d, bg=bg, bold_first=True)
    r = blank(ws, r)

    # ---- Formula ----
    r = section(ws, r, "Composite Risk Score Formula (Line Level)")
    r = blank(ws, r, 4)
    r = formula(ws, r,
                "risk_score  =  0.30 × if_score  +  0.25 × lof_score  +  0.25 × zscore_score"
                "  +  0.15 × rule_flags_score  +  0.05 × benford_score", h=22)
    r = blank(ws, r, 4)
    r = body(ws, r,
             "All component scores are normalised to [0, 1] within the current dataset before "
             "weighting, so 1.0 always represents the most anomalous transaction in the run "
             "for that component. The resulting risk_score is also bounded [0, 1].", h=36)
    r = blank(ws, r)

    # ---- Voucher rollup ----
    r = section(ws, r, "Voucher-Level Rollup Formula")
    r = blank(ws, r, 4)
    r = formula(ws, r,
                "voucher_score  =  0.60 × max_line_score  +  0.25 × mean_line_score"
                "  +  0.15 × flag_density", h=22)
    r = formula(ws, r,
                "  Single-line vouchers:  voucher_score  =  risk_score  (no rollup)", h=16)
    r = blank(ws, r, 4)
    r = body(ws, r,
             "flag_density = total rule flags across all lines ÷ (8 × number of lines). "
             "The 60/25/15 split weights the worst line most heavily, moderates by whether "
             "other lines are also elevated, and adds breadth coverage via flag density.", h=36)


# ===========================================================================
# Sheet 2 — Isolation Forest
# ===========================================================================

def build_if(ws):
    widths(ws, [24, 14, 14, 14, 14, 14, 14, 14])

    r = 1
    _cell(ws, r, 1, "Isolation Forest — Mathematical Reference",
          f=_font(14, bold=True, color=WHITE),
          fl=_fill(NAVY), al=_al("center","center",wrap=False),
          span=SPAN, h=30)
    r += 1

    # ---- Overview ----
    r = section(ws, r, "1. Algorithm Overview")
    r = blank(ws, r, 4)
    r = body(ws, r,
             "Isolation Forest (Liu, Ting & Zhou, 2008) detects anomalies by exploiting the "
             "observation that anomalous points are few and different — they are easier to "
             "isolate from the rest of the data than normal points.", h=36)
    r = body(ws, r,
             "The algorithm builds an ensemble of T = 300 binary isolation trees. Each tree "
             "is grown by repeatedly selecting a random feature and a random split value "
             "(uniform between the feature's min and max in the current subset). Splitting "
             "continues until each point occupies its own leaf or the maximum tree depth is "
             "reached. The number of splits required to isolate a point — its path length "
             "h(x) — is the key quantity: short paths signal anomalies.", h=54)
    r = blank(ws, r)

    # ---- Formulas ----
    r = section(ws, r, "2. Mathematical Formulas")
    r = blank(ws, r, 4)

    r = subsection(ws, r, "2a. Expected Path Length Correction  c(n)")
    r = blank(ws, r, 4)
    r = body(ws, r,
             "The expected path length for an unsuccessful binary search tree (BST) with n "
             "observations is used to normalise path lengths across trees of different sizes:")
    r = formula(ws, r, "c(n)  =  2 · H(n − 1)  −  2·(n − 1) / n")
    r = body(ws, r,
             "where  H(i) = harmonic number = 1 + 1/2 + 1/3 + … + 1/i  ≈  ln(i) + 0.5772")
    r = note(ws, r,
             "Example (n = 1,000 transactions):  H(999) ≈ ln(999) + 0.5772 ≈ 7.473;  "
             "c(1000) = 2 × 7.473 − 2 × 999/1000 ≈ 12.95 − 1.998 ≈ 12.95")
    r = blank(ws, r)

    r = subsection(ws, r, "2b. Per-tree Anomaly Score  s(x, n)  (original paper definition)")
    r = blank(ws, r, 4)
    r = formula(ws, r, "s(x, n)  =  2 ^ ( −E[h(x)] / c(n) )")
    r = body(ws, r,
             "E[h(x)] = average path length of x across all T trees\n"
             "c(n)    = expected path length defined above\n\n"
             "Interpretation:\n"
             "  s → 1.0  :  E[h(x)] << c(n)  →  short paths  →  strongly anomalous\n"
             "  s → 0.5  :  E[h(x)] ≈ c(n)   →  average paths →  cannot distinguish\n"
             "  s → 0.0  :  E[h(x)] >> c(n)  →  long paths   →  definitely normal",
             h=90)
    r = blank(ws, r)

    r = subsection(ws, r, "2c. sklearn score_samples() output")
    r = blank(ws, r, 4)
    r = body(ws, r,
             "sklearn's IsolationForest.score_samples(X) returns a signed raw score per "
             "observation. By convention (sklearn source), the sign is set so that outliers "
             "receive more-negative values:")
    r = formula(ws, r,
                "score_samples(x)  ∝  −E[h(x)]      "
                "# more negative  =  shorter path  =  more anomalous")
    r = body(ws, r,
             "Typical range: approximately −0.5 to 0 for the bulk of inliers, extending "
             "further negative for anomalies. The exact numeric range varies by dataset size "
             "and feature distribution.")
    r = blank(ws, r)

    r = subsection(ws, r, "2d. Normalisation to if_score ∈ [0, 1]")
    r = blank(ws, r, 4)
    r = body(ws, r,
             "Because score_samples() signs anomalies negatively, the sign is flipped before "
             "min-max normalisation:")
    r = formula(ws, r, "raw        =  score_samples(X_scaled)")
    r = formula(ws, r, "flipped    =  −raw                             # anomalous → larger positive")
    r = formula(ws, r,
                "if_score   =  (flipped − min(flipped)) / (max(flipped) − min(flipped))")
    r = body(ws, r,
             "Result:  if_score = 1.0 for the single most anomalous transaction in the run; "
             "if_score = 0.0 for the single least anomalous. All other transactions fall "
             "proportionally between these extremes.", h=42)
    r = blank(ws, r)

    r = subsection(ws, r, "2e. Binary anomaly flag  if_anomaly")
    r = blank(ws, r, 4)
    r = formula(ws, r,
                "if_anomaly  =  1  if  predict(x) == −1,  else  0")
    r = body(ws, r,
             "IsolationForest.predict() uses the internal contamination=0.05 threshold: "
             "the 5% of observations with the lowest score_samples() values are classified "
             "as anomalies (predict = −1). This flag is used in the ML Consensus indicator "
             "and reason codes; it does NOT feed into the composite risk_score formula.")
    r = blank(ws, r)

    # ---- Feature input ----
    r = section(ws, r, "3. Feature Input")
    r = blank(ws, r, 4)
    r = body(ws, r,
             "The Isolation Forest receives the RobustScaler-transformed feature matrix "
             "produced by engineer_features(). RobustScaler centres each feature on its "
             "median and scales by its interquartile range, making the model robust to "
             "the extreme outliers it is trying to detect.")
    r = body(ws, r,
             "Up to 12 features enter the matrix (subject to Spearman correlation pruning "
             "at |corr| > 0.85 within the run):\n"
             "amount_log, amount_zscore_vendor, amount_zscore_costcentre, vendor_txn_count, "
             "processing_days_zscore, desc_length_zscore, vendor_amount_cv, "
             "is_round_number, is_weekend_payment, is_month_end, is_individual_payee, "
             "near_threshold, same_amount_vendor_irregular, is_duplicate, is_reversal",
             h=54)
    r = blank(ws, r)

    # ---- Worked example ----
    r = section(ws, r, "4. Worked Example  (illustrative — simplified to 2 features for clarity)")
    r = blank(ws, r, 4)
    r = body(ws, r,
             "In practice the model operates on ~10 scaled features simultaneously. This "
             "example uses 2 features (Amount z-score, Processing-days z-score) to illustrate "
             "how path lengths map to scores. Path lengths shown are illustrative means across "
             "T=300 trees.")
    r = blank(ws, r, 4)

    r = hdr_row(ws, r, ["Transaction", "Amount z-score", "Processing z-score",
                         "Avg Path Length  E[h(x)]",
                         "score_samples  output",
                         "− score_samples  (flipped)",
                         "if_score  (normalised)",
                         "if_anomaly"])
    ex_data = [
        ("Voucher A  (normal)",  " 0.23",  " 0.41", "9.20", "−0.158", "0.158", "0.047", "0"),
        ("Voucher B  (normal)",  "−0.51",  " 0.18", "9.50", "−0.172", "0.172", "0.140", "0"),
        ("Voucher C  (normal)",  " 0.78",  "−0.33", "8.90", "−0.165", "0.165", "0.093", "0"),
        ("Voucher D  (normal)",  "−0.12",  " 0.56", "9.10", "−0.151", "0.151", "0.000", "0"),
        ("Voucher E  (normal)",  " 0.34",  "−0.08", "9.30", "−0.180", "0.180", "0.193", "0"),
        ("Voucher F  (anomaly)", " 5.82",  " 7.21", "3.10", "−0.301", "0.301", "1.000", "1"),
    ]
    for i, row_d in enumerate(ex_data):
        bg = ALT if i % 2 == 0 else WHITE
        if i == 5:
            bg = L_ORANGE
        r = data_row(ws, r, row_d, bg=bg, bold_first=True)
    r = blank(ws, r, 4)

    r = subsection(ws, r, "Normalisation arithmetic for this example")
    r = blank(ws, r, 4)
    r = formula(ws, r, "min(flipped)  =  0.151   (Voucher D)")
    r = formula(ws, r, "max(flipped)  =  0.301   (Voucher F)")
    r = formula(ws, r, "range         =  0.301 − 0.151  =  0.150")
    r = formula(ws, r, "if_score(A)   =  (0.158 − 0.151) / 0.150  =  0.007 / 0.150  =  0.047")
    r = formula(ws, r, "if_score(D)   =  (0.151 − 0.151) / 0.150  =  0.000")
    r = formula(ws, r, "if_score(F)   =  (0.301 − 0.151) / 0.150  =  1.000")
    r = blank(ws, r, 4)
    r = note(ws, r,
             "Voucher F is flagged if_anomaly = 1 because its score_samples value (−0.301) "
             "falls in the bottom 5% across the full dataset (contamination = 0.05 boundary). "
             "The remaining vouchers are classified as normal (if_anomaly = 0).")


# ===========================================================================
# Sheet 3 — Local Outlier Factor
# ===========================================================================

def build_lof(ws):
    widths(ws, [24, 10, 10, 14, 14, 14, 14, 14])

    r = 1
    _cell(ws, r, 1, "Local Outlier Factor — Mathematical Reference",
          f=_font(14, bold=True, color=WHITE),
          fl=_fill(NAVY), al=_al("center","center",wrap=False),
          span=SPAN, h=30)
    r += 1

    r = section(ws, r, "1. Algorithm Overview")
    r = blank(ws, r, 4)
    r = body(ws, r,
             "LOF (Breunig et al., 2000) measures local density deviation: a point is an "
             "outlier if its neighbourhood density is much lower than that of its k nearest "
             "neighbours. Unlike Isolation Forest, LOF benchmarks each transaction against "
             "its most similar peers rather than the full dataset.", h=42)
    r = body(ws, r,
             "This makes LOF particularly effective at catching inflated invoices — a "
             "SGD 50,000 payment to a vendor whose typical invoices are SGD 2,000 will "
             "score very high even if SGD 50,000 payments exist elsewhere in the dataset.",
             h=36)
    r = blank(ws, r)

    r = section(ws, r, "2. Step-by-Step Mathematical Formulas  (k = n_neighbors = 20 in pipeline)")
    r = blank(ws, r, 4)

    r = subsection(ws, r, "Step 1 — k-Distance of point p")
    r = blank(ws, r, 4)
    r = formula(ws, r,
                "k-dist(p)  =  distance to the k-th nearest neighbour of p")
    r = body(ws, r,
             "In the pipeline, k = 20 and distance is Euclidean in the RobustScaler-"
             "normalised feature space. Points with fewer than k neighbours in the dataset "
             "use min(k, n−1) neighbours.")
    r = blank(ws, r)

    r = subsection(ws, r, "Step 2 — k-Nearest Neighbour Set  N_k(p)")
    r = blank(ws, r, 4)
    r = formula(ws, r,
                "N_k(p)  =  { o ∈ D  |  dist(p, o) ≤ k-dist(p) }")
    r = body(ws, r,
             "The set of all points within k-distance of p (may include ties, so |N_k(p)| ≥ k).")
    r = blank(ws, r)

    r = subsection(ws, r, "Step 3 — Reachability Distance")
    r = blank(ws, r, 4)
    r = formula(ws, r,
                "reach_dist_k(p, o)  =  max( k-dist(o),  dist(p, o) )")
    r = body(ws, r,
             "Smooths short distances: if p is very close to o, the reach distance is "
             "floored at o's k-distance. This prevents artificially high density estimates "
             "for points that happen to be very near a dense cluster centre.")
    r = blank(ws, r)

    r = subsection(ws, r, "Step 4 — Local Reachability Density  lrd_k(p)")
    r = blank(ws, r, 4)
    r = formula(ws, r,
                "lrd_k(p)  =  |N_k(p)|  /  Σ  reach_dist_k(p, o)   for o ∈ N_k(p)")
    r = body(ws, r,
             "= k ÷ (sum of reachability distances from p to each of its k neighbours)\n"
             "High lrd → p's neighbours are close (dense neighbourhood)\n"
             "Low  lrd → p's neighbours are far (sparse neighbourhood)")
    r = blank(ws, r)

    r = subsection(ws, r, "Step 5 — Local Outlier Factor  LOF_k(p)")
    r = blank(ws, r, 4)
    r = formula(ws, r,
                "LOF_k(p)  =  ( Σ  lrd_k(o) / lrd_k(p)   for o ∈ N_k(p) )  /  |N_k(p)|")
    r = body(ws, r,
             "= mean LRD of p's k neighbours ÷ p's own LRD\n\n"
             "LOF ≈ 1  :  p's density matches its neighbourhood  → normal\n"
             "LOF >> 1 :  neighbours are far denser than p       → outlier (sparse island)")
    r = blank(ws, r)

    r = subsection(ws, r, "Step 6 — Normalisation to lof_score ∈ [0, 1]")
    r = blank(ws, r, 4)
    r = formula(ws, r,
                "raw         =  −lof.negative_outlier_factor_    # sklearn stores −LOF (positive)")
    r = formula(ws, r,
                "lof_score   =  (raw − min(raw)) / (max(raw) − min(raw))")
    r = body(ws, r,
             "negative_outlier_factor_ is stored with a negative sign in sklearn. Flipping "
             "and min-max normalising gives lof_score = 1.0 for the transaction most "
             "anomalous relative to its peer group, 0.0 for the least.")
    r = blank(ws, r)

    r = subsection(ws, r, "Step 7 — Binary flag  lof_anomaly")
    r = blank(ws, r, 4)
    r = formula(ws, r,
                "lof_anomaly  =  1  if  fit_predict(x) == −1,  else  0")
    r = body(ws, r,
             "fit_predict() at contamination=0.05 classifies the 5% of observations with "
             "the largest LOF values as anomalies (returns −1). Used in ML Consensus; "
             "does NOT enter the composite risk_score formula.")
    r = blank(ws, r)

    # ---- Worked example ----
    r = section(ws, r, "3. Worked Example  (k = 2, 1-dimensional dataset — illustrative)")
    r = blank(ws, r, 4)
    r = body(ws, r,
             "Dataset: 5 payment amounts on a single dimension after scaling. "
             "k = 2 (simplified from k = 20 in production). "
             "Distances are absolute differences.")
    r = blank(ws, r, 4)

    # Points table
    r = hdr_row(ws, r,
                ["Point", "Scaled Amount", "2-NN (k=2)", "k-dist(p)",
                 "reach_dist to N1", "reach_dist to N2",
                 "lrd_k(p)", "LOF_k(p)"])
    lof_data = [
        ("p1 = 1", "1",  "p2(d=1), p3(d=2)", "2",
         "max(1,1)=1", "max(1,2)=2", "2/(1+2)=0.667", "1.000"),
        ("p2 = 2", "2",  "p1(d=1), p3(d=1)", "1",
         "max(2,1)=2", "max(1,1)=1", "2/(2+1)=0.667", "1.000"),
        ("p3 = 3", "3",  "p2(d=1), p4(d=1)", "1",
         "max(1,1)=1", "max(2,1)=2", "2/(1+2)=0.667", "1.000"),
        ("p4 = 4", "4",  "p3(d=1), p2(d=2)", "2",
         "max(1,1)=1", "max(1,2)=2", "2/(1+2)=0.667", "1.000"),
        ("p5 = 20  ★", "20", "p4(d=16), p3(d=17)", "16",
         "max(2,16)=16", "max(1,17)=17", "2/(16+17)=0.061", "11.000"),
    ]
    for i, row_d in enumerate(lof_data):
        bg = L_ORANGE if i == 4 else (ALT if i % 2 == 0 else WHITE)
        r = data_row(ws, r, row_d, bg=bg, bold_first=True)
    r = blank(ws, r, 4)

    r = subsection(ws, r, "LOF computation detail for p5 (anomaly)")
    r = blank(ws, r, 4)
    r = formula(ws, r,
                "LOF(p5)  =  ( lrd(p4) + lrd(p3) ) / 2  ÷  lrd(p5)")
    r = formula(ws, r,
                "          =  ( 0.667 + 0.667 ) / 2  ÷  0.061")
    r = formula(ws, r,
                "          =  0.667  ÷  0.061  =  11.0")
    r = blank(ws, r, 4)

    r = subsection(ws, r, "Normalisation to lof_score")
    r = blank(ws, r, 4)
    r = formula(ws, r, "min(LOF)  =  1.000   (p1 – p4)")
    r = formula(ws, r, "max(LOF)  =  11.000  (p5)")
    r = formula(ws, r, "lof_score(p1–p4)  =  (1.000 − 1.000) / (11.000 − 1.000)  =  0.000")
    r = formula(ws, r, "lof_score(p5)     =  (11.000 − 1.000) / (11.000 − 1.000)  =  1.000")
    r = blank(ws, r, 4)
    r = note(ws, r,
             "p5 (scaled amount = 20) is a clear outlier: its neighbourhood density is "
             "11× lower than its nearest neighbours' densities. In the payment context "
             "this would represent a transaction far larger than its vendor peer group.")


# ===========================================================================
# Sheet 4 — Benford's Law
# ===========================================================================

def build_benford(ws):
    widths(ws, [22, 12, 14, 14, 14, 14, 14, 12])

    r = 1
    _cell(ws, r, 1, "Benford's Law — Mathematical Reference",
          f=_font(14, bold=True, color=WHITE),
          fl=_fill(NAVY), al=_al("center","center",wrap=False),
          span=SPAN, h=30)
    r += 1

    r = section(ws, r, "1. Benford's Law Principle")
    r = blank(ws, r, 4)
    r = body(ws, r,
             "In naturally occurring financial datasets, the probability that a number "
             "begins with digit d (d ∈ {1, …, 9}) follows Benford's Law:")
    r = formula(ws, r, "P(d)  =  log₁₀( 1 + 1/d )")
    r = body(ws, r,
             "Significant deviation of observed first-digit frequencies from these expected "
             "values may indicate amounts that were manually chosen, rounded, or constructed "
             "rather than arising from genuine transactions.")
    r = blank(ws, r, 4)

    # Expected frequency table
    r = subsection(ws, r, "Expected First-Digit Frequencies")
    r = blank(ws, r, 4)
    r = hdr_row(ws, r,
                ["First Digit  d", "Formula  log₁₀(1 + 1/d)",
                 "Expected  P(d)", "Expected  P(d) %",
                 "Cumulative  P(≤d)", "", "", ""])
    benford = {d: np.log10(1 + 1/d) for d in range(1, 10)}
    cum = 0.0
    for i, d in enumerate(range(1, 10)):
        p = benford[d]
        cum += p
        bg = ALT if i % 2 == 0 else WHITE
        _cell(ws, r, 1, d,          f=_font(9,bold=True), fl=_fill(bg), al=_al("center","center",False), b=TB)
        _cell(ws, r, 2, f"log₁₀(1 + 1/{d})", f=_font(9,mono=True), fl=_fill(bg), al=_al("center","center",False), b=TB)
        _cell(ws, r, 3, round(p, 6),  f=_font(9), fl=_fill(bg), al=_al("center","center",False), b=TB, nf="0.000000")
        _cell(ws, r, 4, f"{p*100:.2f} %", f=_font(9), fl=_fill(bg), al=_al("center","center",False), b=TB)
        _cell(ws, r, 5, f"{cum*100:.2f} %", f=_font(9), fl=_fill(bg), al=_al("center","center",False), b=TB)
        for col in range(6, SPAN+1):
            _cell(ws, r, col, "", fl=_fill(bg), b=TB)
        r += 1
    r = blank(ws, r)

    # ---- Formulas ----
    r = section(ws, r, "2. Score Computation — Four Steps")
    r = blank(ws, r, 4)

    r = subsection(ws, r, "Step 1 — Dataset-Level Analysis  (non-recurring transactions only)")
    r = blank(ws, r, 4)
    r = body(ws, r,
             "Only non-recurring transactions are analysed (is_recurring_payment == 0). "
             "Recurring payments (monthly/quarterly/semi-annual/annual) are excluded because "
             "their fixed amounts naturally deviate from Benford's distribution.")
    r = formula(ws, r,
                "first_digit(amount)  =  first significant digit of abs(amount)  ∈ {1, …, 9}")
    r = formula(ws, r,
                "p_obs(d)  =  count of transactions with first digit d  /  n_nonrecurring")
    r = formula(ws, r,
                "deviation(d)  =  | p_obs(d)  −  P(d) |")
    r = formula(ws, r,
                "MAD  =  mean( deviation(d) )  for d ∈ {1, …, 9}")
    r = formula(ws, r,
                "deviant_digits  =  top-3 digits by deviation(d)  (the most deviant three)")
    r = blank(ws, r)

    r = subsection(ws, r, "Step 2 — Per-Transaction Tagging")
    r = blank(ws, r, 4)
    r = formula(ws, r,
                "benford_flag(tx)  =  1  if  is_recurring_payment == 0"
                "  AND  first_digit(tx) ∈ deviant_digits,  else  0")
    r = formula(ws, r,
                "benford_deviation_score(tx)  =  benford_flag(tx)  ×  deviation( first_digit(tx) )")
    r = body(ws, r,
             "Flagged transactions carry a score equal to the observed deviation magnitude "
             "for their specific first digit. Non-flagged transactions score zero.")
    r = blank(ws, r)

    r = subsection(ws, r, "Step 3 — Normalisation to benford_score ∈ [0, 1]")
    r = blank(ws, r, 4)
    r = formula(ws, r,
                "benford_score(tx)  =  benford_deviation_score(tx)"
                "  /  max( benford_deviation_score )  across all transactions")
    r = body(ws, r,
             "The transaction whose first digit shows the single largest deviation from "
             "Benford's Law receives benford_score = 1.0. All others scale proportionally.")
    r = blank(ws, r)

    r = subsection(ws, r, "Step 4 — Benford Suppression Rule  (applied in sample_selector)")
    r = blank(ws, r, 4)
    r = body(ws, r,
             "A high Benford score alone does not necessarily indicate a suspicious "
             "transaction — naturally noisy datasets can produce deviation. The suppression "
             "rule zeros out the Benford contribution when all other risk signals are weak:")
    r = formula(ws, r,
                "if  if_score < median(if_score)")
    r = formula(ws, r,
                "AND lof_score < median(lof_score)")
    r = formula(ws, r,
                "AND zscore_score < median(zscore_score)")
    r = formula(ws, r,
                "AND rule_flags_score < median(rule_flags_score)")
    r = formula(ws, r,
                "  →  benford_score  =  0.0   (suppressed)")
    r = body(ws, r,
             "Benford evidence is only counted when at least one other signal is also "
             "elevated. This prevents Benford deviation alone from driving sample selection.",
             h=36)
    r = blank(ws, r)

    # ---- Worked example ----
    r = section(ws, r, "3. Worked Example  (20 non-recurring transactions)")
    r = blank(ws, r, 4)
    r = body(ws, r,
             "Hypothetical dataset of 20 non-recurring transactions. Steps 1–3 shown in full.")
    r = blank(ws, r, 4)

    r = subsection(ws, r, "Step 1 — Observed vs Expected Frequencies")
    r = blank(ws, r, 4)
    r = hdr_row(ws, r,
                ["First Digit  d", "Observed Count", "Observed  p_obs(d)",
                 "Expected  P(d)", "Deviation  |p_obs − P(d)|",
                 "Most Deviant?", "", ""])
    ex_obs = {1:8, 2:4, 3:3, 4:1, 5:1, 6:1, 7:1, 8:1, 9:0}
    n_ex = 20
    ex_devs = {}
    for d in range(1, 10):
        p_obs_d = ex_obs[d] / n_ex
        p_exp_d = benford[d]
        ex_devs[d] = abs(p_obs_d - p_exp_d)
    sorted_devs = sorted(ex_devs, key=lambda x: ex_devs[x], reverse=True)
    top3 = set(sorted_devs[:3])
    for i, d in enumerate(range(1, 10)):
        p_obs_d = ex_obs[d] / n_ex
        p_exp_d = benford[d]
        dev_d = ex_devs[d]
        is_top3 = d in top3
        bg = L_ORANGE if is_top3 else (ALT if i % 2 == 0 else WHITE)
        _cell(ws, r, 1, d, f=_font(9,bold=True), fl=_fill(bg), al=_al("center","center",False), b=TB)
        _cell(ws, r, 2, ex_obs[d], f=_font(9), fl=_fill(bg), al=_al("center","center",False), b=TB)
        _cell(ws, r, 3, f"{p_obs_d:.4f}  ({p_obs_d*100:.1f}%)", f=_font(9), fl=_fill(bg), al=_al("center","center",False), b=TB)
        _cell(ws, r, 4, f"{p_exp_d:.4f}  ({p_exp_d*100:.1f}%)", f=_font(9), fl=_fill(bg), al=_al("center","center",False), b=TB)
        _cell(ws, r, 5, round(dev_d, 4), f=_font(9), fl=_fill(bg), al=_al("center","center",False), b=TB, nf="0.0000")
        _cell(ws, r, 6, "★ Top-3 deviant" if is_top3 else "", f=_font(9,bold=is_top3,color=ORANGE if is_top3 else "000000"), fl=_fill(bg), al=_al("center","center",False), b=TB)
        for col in [7, 8]:
            _cell(ws, r, col, "", fl=_fill(bg), b=TB)
        r += 1
    r = blank(ws, r, 4)

    mad_ex = np.mean(list(ex_devs.values()))
    r = formula(ws, r,
                f"MAD  =  mean({', '.join(f'{ex_devs[d]:.4f}' for d in range(1,10))})  =  {mad_ex:.4f}")
    r = formula(ws, r,
                f"deviant_digits  =  {sorted(top3)}  "
                f"(digits with highest absolute deviation from Benford expected)")
    r = blank(ws, r)

    r = subsection(ws, r, "Step 2 & 3 — Per-Transaction Scoring")
    r = blank(ws, r, 4)
    r = hdr_row(ws, r,
                ["Transaction", "Amount (SGD)", "First Digit", "In deviant_digits?",
                 "benford_flag", "deviation score", "benford_score  (normalised)", ""])
    max_dev = max(ex_devs[d] for d in top3)
    ex_tx = [
        ("Tx-01", "SGD 1,250", 1, True),
        ("Tx-02", "SGD 4,800", 4, True),
        ("Tx-03", "SGD 2,300", 2, False),
        ("Tx-04", "SGD 9,100 (recurring)", "—", False),
    ]
    for i, (txid, amt, fd, is_dev) in enumerate(ex_tx):
        if fd == "—":
            flag = 0; dev_s = "0.0000"; norm_s = "0.0000 (recurring — excluded)"
        elif is_dev:
            flag = 1; dev_s = f"{ex_devs[fd]:.4f}"; norm_s = f"{ex_devs[fd]/max_dev:.4f}"
        else:
            flag = 0; dev_s = "0.0000"; norm_s = "0.0000"
        bg = L_ORANGE if (is_dev and fd != "—") else (ALT if i % 2 == 0 else WHITE)
        row_vals = [txid, amt, str(fd), "Yes" if is_dev else "No",
                    str(flag), dev_s, norm_s, ""]
        r = data_row(ws, r, row_vals, bg=bg, bold_first=True)
    r = blank(ws, r, 4)
    r = formula(ws, r,
                f"max(benford_deviation_score)  =  {max_dev:.4f}  (digit {sorted_devs[0]})")
    r = formula(ws, r,
                f"benford_score(Tx-01)  =  {ex_devs[1]:.4f} / {max_dev:.4f}  =  {ex_devs[1]/max_dev:.4f}")
    r = formula(ws, r,
                f"benford_score(Tx-02)  =  {ex_devs[4]:.4f} / {max_dev:.4f}  =  {ex_devs[4]/max_dev:.4f}")
    r = formula(ws, r,
                "benford_score(Tx-03)  =  0.0000  (first digit 2 is not in top-3 deviant digits)")


# ===========================================================================
# Sheet 5 — Composite Score
# ===========================================================================

def build_composite(ws):
    widths(ws, [26, 14, 14, 14, 14, 14, 12, 12])

    r = 1
    _cell(ws, r, 1, "Composite Risk Score — Full Scoring Reference",
          f=_font(14, bold=True, color=WHITE),
          fl=_fill(NAVY), al=_al("center","center",wrap=False),
          span=SPAN, h=30)
    r += 1

    # ---- Line-level formula ----
    r = section(ws, r, "1. Line-Level Composite Risk Score")
    r = blank(ws, r, 4)
    r = formula(ws, r,
                "risk_score  =  0.30 × if_score  +  0.25 × lof_score  +  0.25 × zscore_score"
                "  +  0.15 × rule_flags_score  +  0.05 × benford_score", h=22)
    r = blank(ws, r, 4)

    r = hdr_row(ws, r,
                ["Component", "Weight", "Source Formula", "Score Range",
                 "Anomaly Boundary", "Derivation Sheet", "", ""])
    comp_data = [
        ("if_score", "30%",
         "(−score_samples − min) / (max − min)",
         "[0, 1]", "if_anomaly: predict()==−1 (top 5%)", "Sheet: Isolation Forest", "", ""),
        ("lof_score", "25%",
         "(−neg_outlier_factor − min) / (max − min)",
         "[0, 1]", "lof_anomaly: fit_predict()==−1 (top 5%)", "Sheet: Local Outlier Factor", "", ""),
        ("zscore_score", "25%",
         "(max(|z_vendor|,|z_cc|) − min) / (max − min)",
         "[0, 1]", "zscore_anomaly: max z > 2.0", "Sheet: Composite Score", "", ""),
        ("rule_flags_score", "15%",
         "Σ binary_flags / 8",
         "[0, 1]", "Any rule triggered (score > 0)", "Sheet: Composite Score", "", ""),
        ("benford_score", "5%",
         "deviation(first_digit) / max(deviation)  [suppressed if all others weak]",
         "[0, 1]", "First digit in top-3 deviant digits", "Sheet: Benford's Law", "", ""),
    ]
    for i, row_d in enumerate(comp_data):
        bg = ALT if i % 2 == 0 else WHITE
        r = data_row(ws, r, row_d, bg=bg, bold_first=True)
    r = blank(ws, r)

    # ---- Z-score formula ----
    r = section(ws, r, "2. Statistical Z-Score Component Detail")
    r = blank(ws, r, 4)
    r = body(ws, r,
             "Computed in feature_engineering.py using group statistics per Vendor ID "
             "and per Cost Centre:")
    r = formula(ws, r,
                "amount_zscore_vendor     =  (amount − mean_vendor) / std_vendor")
    r = formula(ws, r,
                "amount_zscore_costcentre =  (amount − mean_cc)     / std_cc")
    r = formula(ws, r,
                "z_max(tx)  =  max( |amount_zscore_vendor|,  |amount_zscore_costcentre| )")
    r = formula(ws, r,
                "zscore_score  =  (z_max − min(z_max)) / (max(z_max) − min(z_max))")
    r = formula(ws, r,
                "zscore_anomaly  =  1  if  z_max > 2.0,  else  0")
    r = body(ws, r,
             "Interpretation: z_max > 2.0 means the amount is more than 2 standard "
             "deviations above the average for that vendor or cost centre — approximately "
             "the upper 2.5% of a normal distribution.")
    r = blank(ws, r)

    # ---- Rule flags ----
    r = section(ws, r, "3. Rule-Based Flags Component Detail")
    r = blank(ws, r, 4)
    r = formula(ws, r,
                "rule_flags_score  =  (Σ flag_i  for i ∈ 8 rules)  /  8")
    r = blank(ws, r, 4)
    r = hdr_row(ws, r, ["Flag Column", "Rule", "Condition", "Value", "", "", "", ""])
    flag_rules = [
        ("is_round_number",             "Round number",             "amount mod 100 == 0",                         "0 or 1"),
        ("is_weekend_payment",          "Weekend payment",          "Invoice Date is Saturday or Sunday",          "0 or 1"),
        ("is_month_end",                "Month-end",                "Invoice day ≥ days_in_month − 2",             "0 or 1"),
        ("near_threshold",              "Near approval threshold",  "amount within 5% below 1K/5K/10K/50K/100K",  "0 or 1"),
        ("is_individual_payee",         "Individual payee",         "Vendor ID matches NRIC/FIN regex",            "0 or 1"),
        ("same_amount_vendor_irregular","Irregular repeat amount",  "Same amount to same vendor >2×, no regular cycle","0 or 1"),
        ("is_duplicate",                "Duplicate payment",        "Same (Vendor ID, Invoice #, Amount) in >1 Voucher ID","0 or 1"),
        ("is_reversal",                 "Reversal / credit note",   "Amount < 0",                                  "0 or 1"),
    ]
    for i, (col, rule, cond, val) in enumerate(flag_rules):
        bg = ALT if i % 2 == 0 else WHITE
        data_row(ws, r, [col, rule, cond, val, "", "", "", ""], bg=bg, bold_first=True)
        r += 1
    r = blank(ws, r)

    # ---- Voucher rollup ----
    r = section(ws, r, "4. Voucher-Level Rollup")
    r = blank(ws, r, 4)
    r = body(ws, r,
             "After line-level risk_scores are computed, lines are grouped by Voucher ID "
             "(the document auditors physically pull):")
    r = formula(ws, r,
                "flag_density   =  total rule flags across all lines  /  (8 × line_count)")
    r = formula(ws, r,
                "voucher_score  =  0.60 × max(risk_score)  +  0.25 × mean(risk_score)"
                "  +  0.15 × flag_density   [multi-line]")
    r = formula(ws, r,
                "voucher_score  =  risk_score   [single-line — no rollup needed]")
    r = blank(ws, r, 4)
    r = body(ws, r,
             "Weight rationale: the 60/25/15 split reflects that audit significance is "
             "primarily driven by the worst line in the voucher, moderated by whether other "
             "lines are also elevated, and supplemented by the breadth of rule flag coverage.")
    r = blank(ws, r)

    # ---- Risk tiers ----
    r = section(ws, r, "5. Risk Tier Assignment")
    r = blank(ws, r, 4)
    r = formula(ws, r,
                "HIGH    :  voucher_score ≥ 95th percentile of all vouchers in run")
    r = formula(ws, r,
                "MEDIUM  :  80th ≤ voucher_score < 95th percentile")
    r = formula(ws, r,
                "LOW     :  voucher_score < 80th percentile")
    r = blank(ws, r, 4)
    r = body(ws, r,
             "Percentile-based tiers adapt to any dataset size and composition. HIGH always "
             "covers the most anomalous 5% regardless of absolute score values. All HIGH "
             "vouchers are mandatory in the audit sample; remaining slots are filled "
             "proportionally from MEDIUM (~75%) and LOW (baseline).")

    # ---- ML Consensus ----
    r = blank(ws, r)
    r = section(ws, r, "6. ML Consensus Flag")
    r = blank(ws, r, 4)
    r = body(ws, r,
             "A supplementary display indicator — does NOT alter risk_score or voucher_score:")
    r = formula(ws, r,
                "ML_Consensus_Flag(tx)  =  if_anomaly  +  lof_anomaly  +  zscore_anomaly   [0, 1, 2, or 3]")
    r = formula(ws, r,
                "voucher_any_ml_consensus  =  1  if  any line in voucher has ML_Consensus_Flag ≥ 2")
    r = body(ws, r,
             "When 2 or more independent ML methods agree a transaction is anomalous, the "
             "probability of a true anomaly is materially higher than when only one flags it.")


# ===========================================================================
# Entry point
# ===========================================================================

def main():
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Overview"
    build_overview(ws1)

    ws2 = wb.create_sheet("Isolation Forest")
    build_if(ws2)

    ws3 = wb.create_sheet("Local Outlier Factor")
    build_lof(ws3)

    ws4 = wb.create_sheet("Benford's Law")
    build_benford(ws4)

    ws5 = wb.create_sheet("Composite Score")
    build_composite(ws5)

    out = Path("output/Scoring_Methodology.xlsx")
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
