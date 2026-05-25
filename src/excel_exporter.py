from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

AMOUNT_COL = 'Payment Voucher Amount (SGD, Excluding GST)'

# Colour palette
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
HIGH_FILL    = PatternFill("solid", fgColor="FFB3B3")   # light red
MED_FILL     = PatternFill("solid", fgColor="FFDDB3")   # light orange
LOW_FILL     = PatternFill("solid", fgColor="FFFAB3")   # light yellow
ALT_FILL     = PatternFill("solid", fgColor="F2F2F2")   # light grey
ALT2_FILL    = PatternFill("solid", fgColor="E8EFF8")   # soft blue-grey (voucher alternation)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")   # soft blue

THIN = Side(style="thin", color="BBBBBB")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

_TIER_FILL = {'HIGH': HIGH_FILL, 'MEDIUM': MED_FILL, 'LOW': LOW_FILL}


def _auto_width(ws, min_w=8, max_w=50):
    for col_cells in ws.columns:
        width = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(width + 2, min_w), max_w)


def _write_header_row(ws, headers, row=1):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def _safe_value(value):
    """Convert numpy types and NaN for Excel compatibility."""
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return None if np.isnan(value) else float(value)
    if isinstance(value, float) and np.isnan(value):
        return None
    return value


def _amber_note(ws, text, n_cols, row=1):
    """Full-width amber background note spanning n_cols, used atop review sheets."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = PatternFill("solid", fgColor="FFC000")
    cell.font = Font(size=10)
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)


def _write_summary_block(ws, start_row, items, label_span=3):
    """Write (label, value) rows below a data table: navy-header labels, white-background values."""
    for i, (label, value) in enumerate(items):
        r = start_row + i
        lc = ws.cell(row=r, column=1, value=label)
        lc.fill = HEADER_FILL
        lc.font = HEADER_FONT
        lc.alignment = Alignment(vertical='center', wrap_text=True)
        if label_span > 1:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=label_span)
        vc = ws.cell(row=r, column=label_span + 1, value=value)
        vc.font = Font(size=10)
        vc.fill = PatternFill("solid", fgColor="FFFFFF")
    return start_row + len(items)


# ---------------------------------------------------------------------------
# Sheet 1 — Selected Vouchers
# ---------------------------------------------------------------------------

_VCH_DISPLAY_COLS = [
    'Voucher ID', 'Vendor ID', 'Vendor Name', 'Invoice Number(s)',
    'Voucher Line Description(s)',
    'voucher_total_amount', 'voucher_line_count', 'voucher_score', 'voucher_risk_tier',
    'voucher_flag_count', 'voucher_any_ml_consensus', 'vendor_capped',
    'voucher_reason_codes',
]

_VCH_HEADERS = [
    'Voucher ID', 'Vendor ID', 'Vendor Name', 'Invoice Number(s)',
    'Voucher Line Description(s)',
    'Total Amount (SGD)', 'Line Count', 'Voucher Score', 'Risk Tier',
    'Flag Count', 'ML Consensus', 'Vendor Capped',
    'Reason Codes',
]


def _sheet_selected_vouchers(wb, selected_vouchers):
    ws = wb.active
    ws.title = "Selected Vouchers"
    ws.freeze_panes = "B2"

    headers = ['Sample #'] + _VCH_HEADERS
    _write_header_row(ws, headers)

    cols = ['Sample #'] + _VCH_DISPLAY_COLS
    present = [c for c in cols if c in selected_vouchers.columns]

    for r_idx, row_data in enumerate(
            selected_vouchers[present].itertuples(index=False), start=2):
        tier_idx = present.index('voucher_risk_tier') if 'voucher_risk_tier' in present else -1
        tier = row_data[tier_idx] if tier_idx >= 0 else 'LOW'
        row_fill = _TIER_FILL.get(tier, LOW_FILL)

        for c_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = _safe_value(value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.fill = row_fill

            col_name = present[c_idx - 1]
            if col_name == 'voucher_score':
                cell.number_format = '0.0000'
            elif col_name == 'voucher_total_amount':
                cell.number_format = '#,##0.00'

    ws.row_dimensions[1].height = 30
    _auto_width(ws)
    n = len(headers)
    ws.column_dimensions[get_column_letter(n)].width = 60       # Reason Codes
    # Fix width for Voucher Line Description(s) column
    if 'Voucher Line Description(s)' in headers:
        desc_col = headers.index('Voucher Line Description(s)') + 1
        ws.column_dimensions[get_column_letter(desc_col)].width = 50


# ---------------------------------------------------------------------------
# Sheet 2 — Voucher Line Detail
# ---------------------------------------------------------------------------

_ORIG_COLS = [
    'Vendor Name', 'Vendor ID', 'Cost Centre', 'Account Code',
    'Invoice Date', 'Voucher Accounting Date',
    'Invoice Number', 'Voucher ID', 'Voucher Line Description',
    AMOUNT_COL,
]

_LINE_SCORE_COLS = {
    'risk_score':        'Risk Score',
    'if_score':          'Isolation Forest',
    'lof_score':         'Local Outlier',
    'zscore_score':      'Z-Score Signal',
    'benford_score':     "Benford Score",
    'rule_flags_score':  'Rule Flags Score',
    'ML_Consensus_Flag': 'ML Consensus Count',
}

_LINE_FLAG_COLS = [
    'is_round_number', 'is_weekend_payment',
    'near_threshold', 'is_individual_payee',
    'same_amount_vendor_irregular', 'is_duplicate', 'is_reversal',
    'is_split_purchase_risk', 'is_transposed_amount',
    'is_recurring_payment', 'benford_flag', 'processing_days',
    'if_anomaly', 'lof_anomaly', 'zscore_anomaly',
]


def _sheet_voucher_line_detail(wb, df_scored, selected_vouchers):
    ws = wb.create_sheet("Voucher Line Detail")
    ws.freeze_panes = "B2"

    selected_vids = set(selected_vouchers['Voucher ID']) \
        if 'Voucher ID' in selected_vouchers.columns else set()

    if selected_vids and 'Voucher ID' in df_scored.columns:
        lines = df_scored[df_scored['Voucher ID'].isin(selected_vids)].copy()
    else:
        lines = df_scored.copy()

    lines = lines.sort_values(
        ['Voucher ID', 'risk_score'], ascending=[True, False]
    ).reset_index(drop=True)

    orig_present  = [c for c in _ORIG_COLS if c in lines.columns]
    score_present = [c for c in _LINE_SCORE_COLS if c in lines.columns]
    flag_present  = [c for c in _LINE_FLAG_COLS if c in lines.columns]
    reason_col    = ['_line_reason'] if '_line_reason' in lines.columns else []

    display_cols = orig_present + score_present + flag_present + reason_col
    rename_map = {**_LINE_SCORE_COLS, '_line_reason': 'Line Reason Codes'}
    sub = lines[display_cols].rename(columns=rename_map)

    _write_header_row(ws, list(sub.columns))

    date_fmt = {'Invoice Date': 'DD/MM/YYYY', 'Voucher Accounting Date': 'DD/MM/YYYY'}

    # Alternate shading by Voucher ID group
    fills = [ALT_FILL, ALT2_FILL]
    fill_idx = 0
    prev_vid = None

    for r_idx, orig_row in lines[display_cols].iterrows():
        cur_vid = orig_row.get('Voucher ID', None) if 'Voucher ID' in lines.columns else None
        if cur_vid != prev_vid:
            fill_idx = 1 - fill_idx
            prev_vid = cur_vid
        row_fill = fills[fill_idx]

        for c_idx, (col_name, value) in enumerate(
                zip(sub.columns, orig_row[display_cols].values), start=1):
            cell = ws.cell(row=r_idx + 2, column=c_idx)
            cell.value = _safe_value(value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=False)
            cell.fill = row_fill
            if col_name in date_fmt:
                cell.number_format = date_fmt[col_name]

    ws.row_dimensions[1].height = 30
    _auto_width(ws)
    if 'Line Reason Codes' in sub.columns:
        reason_idx = list(sub.columns).index('Line Reason Codes') + 1
        ws.column_dimensions[get_column_letter(reason_idx)].width = 55


# ---------------------------------------------------------------------------
# Sheet 3 — All Vouchers Scored
# ---------------------------------------------------------------------------

_ALL_VCH_COLS = [
    'Voucher ID', 'Vendor ID', 'Vendor Name', 'Invoice Number(s)',
    'voucher_total_amount', 'voucher_line_count', 'voucher_score', 'voucher_risk_tier',
    'voucher_max_score', 'voucher_mean_score',
    'voucher_flag_count', 'voucher_any_ml_consensus',
    'voucher_reason_codes',
]

_ALL_VCH_HEADERS = [
    'Voucher ID', 'Vendor ID', 'Vendor Name', 'Invoice Number(s)',
    'Total Amount (SGD)', 'Line Count', 'Voucher Score', 'Risk Tier',
    'Max Line Score', 'Mean Line Score',
    'Flag Count', 'ML Consensus', 'Reason Codes',
]


def _sheet_all_vouchers(wb, df_vouchers):
    ws = wb.create_sheet("All Vouchers Scored")
    ws.freeze_panes = "B2"

    present = [c for c in _ALL_VCH_COLS if c in df_vouchers.columns]
    headers = [_ALL_VCH_HEADERS[_ALL_VCH_COLS.index(c)] for c in present]
    _write_header_row(ws, headers)

    for r_idx, row_data in enumerate(
            df_vouchers[present].itertuples(index=False), start=2):
        fill = ALT_FILL if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        for c_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = _safe_value(value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=False)
            cell.fill = fill

            col_name = present[c_idx - 1]
            if col_name in ('voucher_score', 'voucher_max_score', 'voucher_mean_score'):
                cell.number_format = '0.0000'
            elif col_name == 'voucher_total_amount':
                cell.number_format = '#,##0.00'

    ws.row_dimensions[1].height = 30
    _auto_width(ws)

    if 'voucher_score' in present:
        sc_letter = get_column_letter(present.index('voucher_score') + 1)
        ws.conditional_formatting.add(
            f"{sc_letter}2:{sc_letter}{len(df_vouchers) + 1}",
            ColorScaleRule(
                start_type='min', start_color='63BE7B',
                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                end_type='max', end_color='F8696B',
            )
        )

    if 'voucher_reason_codes' in present:
        rc_letter = get_column_letter(present.index('voucher_reason_codes') + 1)
        ws.column_dimensions[rc_letter].width = 60


# ---------------------------------------------------------------------------
# Sheet 4 — All Lines Scored  (full row-level dataset, reference)
# ---------------------------------------------------------------------------

_SCORE_COLS_DISPLAY = {
    'risk_score':       'Risk Score',
    'if_score':         'Isolation Forest Score',
    'lof_score':        'Local Outlier Score',
    'zscore_score':     'Z-Score Signal',
    'benford_score':    "Benford's Score",
    'rule_flags_score': 'Rule Flags Score',
}


def _sheet_all_lines(wb, df_scored):
    ws = wb.create_sheet("All Lines Scored")
    ws.freeze_panes = "B2"

    score_cols = [c for c in _SCORE_COLS_DISPLAY if c in df_scored.columns]
    flag_cols  = [c for c in [
        'is_round_number', 'is_weekend_payment',
        'near_threshold', 'is_individual_payee',
        'same_amount_vendor_irregular', 'is_duplicate', 'is_reversal',
        'is_split_purchase_risk', 'is_transposed_amount',
        'is_recurring_payment', 'benford_flag', 'processing_days',
        'if_anomaly', 'lof_anomaly', 'zscore_anomaly',
    ] if c in df_scored.columns]

    orig_present = [c for c in _ORIG_COLS if c in df_scored.columns]
    display_cols = orig_present + score_cols + flag_cols
    score_rename = {k: v for k, v in _SCORE_COLS_DISPLAY.items() if k in score_cols}
    sub = df_scored[display_cols].rename(columns=score_rename)

    _write_header_row(ws, list(sub.columns))

    date_fmt = {'Invoice Date': 'DD/MM/YYYY', 'Voucher Accounting Date': 'DD/MM/YYYY'}

    for r_idx, row_data in enumerate(sub.itertuples(index=False), start=2):
        fill = ALT_FILL if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for c_idx, (col_name, value) in enumerate(zip(sub.columns, row_data), start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = _safe_value(value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=False)
            cell.fill = fill
            if col_name in date_fmt:
                cell.number_format = date_fmt[col_name]

    ws.row_dimensions[1].height = 30
    _auto_width(ws)

    risk_col_idx = next(
        (i + 1 for i, c in enumerate(sub.columns) if c == 'Risk Score'), None
    )
    if risk_col_idx:
        rc = get_column_letter(risk_col_idx)
        ws.conditional_formatting.add(
            f"{rc}2:{rc}{len(sub) + 1}",
            ColorScaleRule(
                start_type='min', start_color='63BE7B',
                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                end_type='max', end_color='F8696B',
            )
        )


# ---------------------------------------------------------------------------
# Sheet — Reversals Review
# ---------------------------------------------------------------------------

_REVERSALS_NOTE = (
    "Reversals Review — All reversal and credit note transactions (negative amounts). "
    "Reversals are a normal accounting mechanism but can also be used to manipulate or disguise "
    "payments. The matched original payment column shows the likely counterpart positive payment "
    "where one could be identified by matching vendor and absolute amount. Auditors should "
    "review whether each reversal is supported and appropriate."
)

_REVERSALS_HEADERS = [
    'Vendor ID', 'Vendor Name', 'Voucher ID', 'Invoice Number',
    'Invoice Date', 'Voucher Accounting Date', 'Voucher Line Description',
    'Amount (SGD)', 'Matched Original Payment (Voucher ID)',
]


def _sheet_reversals_review(wb, df_scored):
    ws = wb.create_sheet("Reversals Review")
    headers = _REVERSALS_HEADERS
    n_cols = len(headers)

    if 'is_reversal' in df_scored.columns:
        rev = df_scored[df_scored['is_reversal'] == 1].copy()
    else:
        rev = df_scored.iloc[0:0].copy()

    if rev.empty:
        _write_header_row(ws, headers, row=1)
        ws.cell(row=2, column=1,
                value="No reversal or credit note transactions identified in this dataset.")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        _auto_width(ws)
        return

    # Lookup of positive payments: (Vendor ID, rounded amount) -> distinct Voucher IDs
    pos = df_scored[df_scored[AMOUNT_COL] > 0]
    pos_lookup = {}
    for vid_, amt_, vch_ in zip(pos['Vendor ID'], pos[AMOUNT_COL], pos['Voucher ID']):
        if pd.isna(amt_):
            continue
        pos_lookup.setdefault((vid_, round(float(amt_), 2)), set()).add(str(vch_))

    def _matched(vid_, amt_):
        if pd.isna(amt_):
            return "(no matching original payment identified)"
        vids = pos_lookup.get((vid_, round(abs(float(amt_)), 2)))
        if not vids:
            return "(no matching original payment identified)"
        return ", ".join(sorted(vids))

    rev = rev.sort_values(['Vendor ID', 'Voucher Accounting Date']).reset_index(drop=True)

    _amber_note(ws, _REVERSALS_NOTE, n_cols, row=1)
    ws.row_dimensions[1].height = 60
    _write_header_row(ws, headers, row=2)

    date_cols = {'Invoice Date', 'Voucher Accounting Date'}
    fills = [ALT_FILL, ALT2_FILL]
    fill_idx = 0
    prev_vid = None
    n_no_match = 0

    for i, (_, r) in enumerate(rev.iterrows()):
        cur_vid = r.get('Vendor ID')
        if cur_vid != prev_vid:
            fill_idx = 1 - fill_idx
            prev_vid = cur_vid
        row_fill = fills[fill_idx]
        matched = _matched(cur_vid, r.get(AMOUNT_COL))
        if matched.startswith("(no matching"):
            n_no_match += 1

        values = [
            r.get('Vendor ID', ''), r.get('Vendor Name', ''), r.get('Voucher ID', ''),
            r.get('Invoice Number', ''), r.get('Invoice Date', None),
            r.get('Voucher Accounting Date', None), r.get('Voucher Line Description', ''),
            r.get(AMOUNT_COL, None), matched,
        ]
        excel_row = i + 3
        for c_idx, (hdr, value) in enumerate(zip(headers, values), start=1):
            cell = ws.cell(row=excel_row, column=c_idx)
            cell.value = _safe_value(value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=False)
            cell.fill = row_fill
            if hdr in date_cols:
                cell.number_format = 'DD/MM/YYYY'
            elif hdr == 'Amount (SGD)':
                cell.number_format = '#,##0.00'

    ws.freeze_panes = "A3"
    _auto_width(ws)
    ws.column_dimensions[get_column_letter(headers.index('Voucher Line Description') + 1)].width = 40
    ws.column_dimensions[get_column_letter(n_cols)].width = 30

    n_rev = len(rev)
    n_vendors = rev['Vendor ID'].nunique()
    total_abs = float(rev[AMOUNT_COL].abs().sum())
    summary_start = n_rev + 4  # note(1) + header(2) + n_rev data rows + 1 blank
    _write_summary_block(ws, summary_start, [
        ("Total reversal transactions", f"{n_rev:,}"),
        ("Distinct vendors with reversals", f"{n_vendors:,}"),
        ("Total absolute value (SGD)", f"{total_abs:,.2f}"),
        ("Reversals with no matching original payment", f"{n_no_match:,}"),
    ])


# ---------------------------------------------------------------------------
# Sheet — FY Split Purchase
# ---------------------------------------------------------------------------

_FY_SPLIT_NOTE = (
    "Potential FY Split Purchase — Payments to the same vendor with similar descriptions within "
    "the same fiscal year (1 Apr - 31 Mar), where the combined total exceeds SGD 6,000. This may "
    "indicate procurement splitting to avoid the small value purchase approval threshold. Fiscal "
    "year is determined by Voucher Accounting Date. Where a payment's description is a unique "
    "reference code rather than descriptive text, it can only be grouped with other payments "
    "carrying the identical reference; vendors whose payments each use a different reference code "
    "will not be grouped and should be reviewed separately. This feature is a review aid only and "
    "does not influence the risk score."
)

_FY_SPLIT_SOURCE_COLS = [
    'Vendor ID', 'Vendor Name', 'fy_split_fy_label', 'Voucher ID', 'Invoice Number',
    'Invoice Date', 'Voucher Accounting Date', 'Voucher Line Description',
    AMOUNT_COL, 'fy_split_group_count', 'fy_split_group_total', 'Account Code', 'Cost Centre',
]

_FY_SPLIT_HEADERS = [
    'Vendor ID', 'Vendor Name', 'Fiscal Year', 'Voucher ID', 'Invoice Number',
    'Invoice Date', 'Voucher Accounting Date', 'Voucher Line Description',
    'Amount (SGD)', 'No. of Similar Payments in FY', 'Group Total (SGD)',
    'Account Code', 'Cost Centre',
]


def _sheet_fy_split_purchase(wb, df_scored):
    ws = wb.create_sheet("FY Split Purchase")
    headers = _FY_SPLIT_HEADERS
    n_cols = len(headers)

    if 'is_fy_split_purchase' in df_scored.columns:
        fy = df_scored[df_scored['is_fy_split_purchase'] == 1].copy()
    else:
        fy = df_scored.iloc[0:0].copy()

    if fy.empty:
        _write_header_row(ws, headers, row=1)
        ws.cell(row=2, column=1,
                value="No potential FY split purchases identified in this dataset.")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        _auto_width(ws)
        return

    fy = fy.sort_values(
        ['Vendor ID', 'fy_split_fy_label', 'Voucher Accounting Date']
    ).reset_index(drop=True)

    _amber_note(ws, _FY_SPLIT_NOTE, n_cols, row=1)
    ws.row_dimensions[1].height = 75
    _write_header_row(ws, headers, row=2)

    date_cols = {'Invoice Date', 'Voucher Accounting Date'}
    amt_cols  = {'Amount (SGD)', 'Group Total (SGD)'}
    fills = [ALT_FILL, ALT2_FILL]
    fill_idx = 0
    prev_vid = None

    for i, (_, r) in enumerate(fy.iterrows()):
        cur_vid = r.get('Vendor ID')
        if cur_vid != prev_vid:
            fill_idx = 1 - fill_idx
            prev_vid = cur_vid
        row_fill = fills[fill_idx]
        excel_row = i + 3
        for c_idx, (src, hdr) in enumerate(zip(_FY_SPLIT_SOURCE_COLS, headers), start=1):
            cell = ws.cell(row=excel_row, column=c_idx)
            cell.value = _safe_value(r.get(src, None))
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=False)
            cell.fill = row_fill
            if hdr in date_cols:
                cell.number_format = 'DD/MM/YYYY'
            elif hdr in amt_cols:
                cell.number_format = '#,##0.00'

    ws.freeze_panes = "A3"
    _auto_width(ws)
    ws.column_dimensions[get_column_letter(headers.index('Voucher Line Description') + 1)].width = 40

    n_flagged = len(fy)
    n_vendors = fy['Vendor ID'].nunique()
    n_groups  = fy[['Vendor ID', 'fy_split_fy_label', 'fy_split_group_total']] \
        .drop_duplicates().shape[0]
    total_sgd = float(fy[AMOUNT_COL].sum())
    summary_start = n_flagged + 4  # note(1) + header(2) + data rows + 1 blank
    _write_summary_block(ws, summary_start, [
        ("Total flagged transactions", f"{n_flagged:,}"),
        ("Distinct vendors affected", f"{n_vendors:,}"),
        ("Distinct vendor-FY-description groups", f"{n_groups:,}"),
        ("Total value across flagged transactions (SGD)", f"{total_sgd:,.2f}"),
    ])


# ---------------------------------------------------------------------------
# Sheet 5 — Benford's Law (unchanged)
# ---------------------------------------------------------------------------

def _sheet_benford(wb, benford_summary, stats):
    ws = wb.create_sheet("Benford's Law")

    ws['A1'] = "Benford's Law Analysis"
    ws['A1'].font = Font(bold=True, size=14, color="1F3864")
    ws.merge_cells('A1:G1')

    ws['A2'] = (
        f"Analysed {stats['n_analyzed']:,} non-recurring transactions  |  "
        f"Excluded {stats['n_excluded_recurring']:,} recurring payments"
    )
    ws['A2'].font = Font(italic=True, size=10, color="444444")
    ws.merge_cells('A2:G2')
    ws.row_dimensions[2].height = 18

    stats_data = [
        ("MAD (Mean Absolute Deviation)", f"{stats['mad']:.4f}"),
        ("Conformity Verdict", stats['conformity']),
        ("Chi-Square Statistic", f"{stats['chi2']:.4f}"),
        ("Chi-Square p-value", f"{stats['p_value']:.4f}"),
        ("Most Deviant Digits", ", ".join(str(d) for d in stats['deviant_digits'])),
    ]

    for row_offset, (label, value) in enumerate(stats_data, start=4):
        ws.cell(row=row_offset, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_offset, column=2, value=value)

    verdict_cell = ws.cell(row=5, column=2)
    conformity_colors = {
        "Conformity": "00B050",
        "Acceptable": "70AD47",
        "Marginally Acceptable": "FFC000",
        "Non-Conformity": "FF0000",
    }
    c = conformity_colors.get(stats['conformity'], "000000")
    verdict_cell.font = Font(bold=True, color=c)

    # Digit frequency table — placed immediately after stats
    tbl_start = 10
    _write_header_row(ws, list(benford_summary.columns), row=tbl_start)
    for r_idx, row_data in enumerate(benford_summary.itertuples(index=False), start=tbl_start + 1):
        digit = row_data[0]
        is_deviant = digit in stats['deviant_digits']
        for c_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
            if is_deviant:
                cell.fill = PatternFill("solid", fgColor="FFE0CC")
    ws.row_dimensions[tbl_start].height = 25

    note_row = tbl_start + len(benford_summary) + 2
    ws.cell(row=note_row, column=1,
            value="Note: Recurring payments (monthly, quarterly, semi-annual, annual) "
                  "are excluded from this analysis as they naturally deviate from "
                  "Benford's distribution without being suspicious.").font = Font(italic=True, size=9, color="666666")
    ws.merge_cells(f'A{note_row}:G{note_row}')

    # Explanation block for MAD, Chi-Square, and Conformity Verdict
    expl_hdr = note_row + 2
    ws.cell(row=expl_hdr, column=1,
            value="Understanding These Metrics").font = Font(bold=True, size=11, color="1F3864")
    ws.cell(row=expl_hdr, column=1).fill = PatternFill("solid", fgColor="D9E1F2")
    ws.merge_cells(f'A{expl_hdr}:G{expl_hdr}')
    ws.row_dimensions[expl_hdr].height = 18

    _explanations = [
        (expl_hdr + 1, "MAD (Mean Absolute Deviation)",
         "Measures the average absolute difference between observed and Benford-expected first-digit "
         "frequencies. Thresholds (Nigrini, 2012): < 0.006 = Close Conformity; "
         "0.006–0.012 = Acceptable Conformity; 0.012–0.015 = Marginally Acceptable; "
         "> 0.015 = Non-Conformity. A lower MAD means the data more closely follows Benford's Law. "
         "MAD is the primary practical measure for audit interpretation."),
        (expl_hdr + 2, "Chi-Square Statistic & p-value",
         "Tests whether the observed digit frequencies are statistically significantly different from "
         "Benford's expected values. A p-value < 0.05 indicates the difference is statistically "
         "significant. A significant chi-square p-value with small MAD (< 0.012) indicates that "
         "anomalies are not pervasive at the overall dataset level, but may be concentrated in "
         "specific transactions or digit groups. In such cases, individual transaction Benford flags "
         "remain relevant and should be reviewed in conjunction with other risk signals."),
        (expl_hdr + 3, "Conformity Verdict",
         "Summarises the overall finding based on the MAD threshold. Non-Conformity does not mean "
         "fraud — it means the first-digit distribution is unusual and warrants investigation of the "
         "most deviant digits. The tool assigns Benford's Law only a 5% weight in the composite risk "
         "score and further suppresses it when all other risk signals are below average, so a "
         "Non-Conformity verdict will not on its own cause any voucher to be selected for audit."),
    ]
    for _rn, _lbl, _txt in _explanations:
        ws.cell(row=_rn, column=1, value=_lbl).font = Font(bold=True, size=10, color="1F3864")
        _ec = ws.cell(row=_rn, column=2, value=_txt)
        _ec.font = Font(size=10)
        _ec.alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'B{_rn}:G{_rn}')
        ws.row_dimensions[_rn].height = 60

    # Key Takeaway — section header + dynamic interpretation for this dataset
    _kt_hdr_row = expl_hdr + 4
    _kt_hdr = ws.cell(row=_kt_hdr_row, column=1, value="Key Takeaway")
    _kt_hdr.font = Font(bold=True, size=11, color="1F3864")
    _kt_hdr.fill = PatternFill("solid", fgColor="D9E1F2")
    _kt_hdr.alignment = Alignment(vertical='center')
    ws.merge_cells(f'A{_kt_hdr_row}:G{_kt_hdr_row}')
    ws.row_dimensions[_kt_hdr_row].height = 18

    _p_sig = stats['p_value'] < 0.05
    _mad_high = stats['mad'] > 0.015
    _mad = stats['mad']
    _pval = stats['p_value']
    if not _p_sig:
        _key_text = (
            f"MAD of {_mad:.4f} with a non-significant p-value (p = {_pval:.4f}) indicates no "
            f"statistically significant deviation from Benford's expected distribution has been "
            f"detected at this dataset size. Benford's Law signals are advisory only; the composite "
            f"risk score is driven primarily by other components (Isolation Forest, Local Outlier "
            f"Factor, z-score, rule-based flags). The absence of a significant result does not "
            f"confirm data integrity — it means no population-level distributional anomaly has been "
            f"detected at this sample size."
        )
    elif _mad_high:
        _key_text = (
            f"MAD of {_mad:.4f} with a statistically significant p-value (p = {_pval:.4f}) "
            f"indicates the distortion is large enough to be visible at the aggregate level. This "
            f"suggests anomalies are either widespread across many transactions, or concentrated "
            f"transactions are extreme enough to visibly drag the overall distribution. A broader "
            f"review of the dataset is warranted — not just a focus on a few deviant digit groups."
        )
    else:
        _key_text = (
            f"MAD of {_mad:.4f} with a statistically significant p-value (p = {_pval:.4f}) "
            f"indicates the overall distribution still looks broadly healthy, but the anomaly is "
            f"real and subtle. This suggests fewer transactions are involved, or any manipulation "
            f"is more targeted. The audit response should be more surgical: focus on patterns "
            f"within the flagged digit groups rather than the dataset as a whole."
        )
    _kt_body_row = _kt_hdr_row + 1
    _kc = ws.cell(row=_kt_body_row, column=1, value=_key_text)
    _kc.font = Font(size=10, color="1F3864")
    _kc.alignment = Alignment(wrap_text=True, vertical='top')
    _kc.fill = PatternFill("solid", fgColor="FFF2CC")
    ws.merge_cells(f'A{_kt_body_row}:G{_kt_body_row}')
    ws.row_dimensions[_kt_body_row].height = 80

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 6 — Summary
# ---------------------------------------------------------------------------

def _sheet_summary(wb, df_scored, df_vouchers, selected_vouchers, benford_stats):
    ws = wb.create_sheet("Summary")

    ws['A1'] = "Payment Audit — Summary"
    ws['A1'].font = Font(bold=True, size=14, color="1F3864")
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 28

    n_lines    = len(df_scored)
    n_vouchers = len(df_vouchers)
    avg_lines  = n_lines / n_vouchers if n_vouchers > 0 else 0
    n_sel      = len(selected_vouchers)
    n_sel_high = int((selected_vouchers.get('voucher_risk_tier', pd.Series()) == 'HIGH').sum())
    n_sel_med  = int((selected_vouchers.get('voucher_risk_tier', pd.Series()) == 'MEDIUM').sum())
    n_sel_low  = int((selected_vouchers.get('voucher_risk_tier', pd.Series()) == 'LOW').sum())
    n_vch_high = int((df_vouchers.get('voucher_risk_tier', pd.Series()) == 'HIGH').sum())
    n_vch_med  = int((df_vouchers.get('voucher_risk_tier', pd.Series()) == 'MEDIUM').sum())
    n_vch_low  = int((df_vouchers.get('voucher_risk_tier', pd.Series()) == 'LOW').sum())

    rev_mask    = df_scored.get('is_reversal', pd.Series(0, index=df_scored.index)) == 1
    n_reversals = int(rev_mask.sum())
    abs_rev_tot = float(df_scored.loc[rev_mask, AMOUNT_COL].abs().sum()) if n_reversals else 0.0

    fy_mask      = df_scored.get('is_fy_split_purchase', pd.Series(0, index=df_scored.index)) == 1
    n_fy_flagged = int(fy_mask.sum())
    if n_fy_flagged and {'fy_split_fy_label', 'fy_split_group_total'} <= set(df_scored.columns):
        n_fy_groups = df_scored.loc[
            fy_mask, ['Vendor ID', 'fy_split_fy_label', 'fy_split_group_total']
        ].drop_duplicates().shape[0]
        fy_total = float(df_scored.loc[fy_mask, AMOUNT_COL].sum())
    else:
        n_fy_groups, fy_total = 0, 0.0

    rows = [
        ("DATASET", None),
        ("Total transaction line items", f"{n_lines:,}"),
        ("Unique payment vouchers", f"{n_vouchers:,}"),
        ("Average lines per voucher", f"{avg_lines:.1f}"),
        ("Recurring payments excluded from Benford's", f"{benford_stats.get('n_excluded_recurring', 0):,}"),
        ("", None),
        ("VOUCHER RISK TIERS (all vouchers)", None),
        ("HIGH risk vouchers (top 5%)", f"{n_vch_high:,}"),
        ("MEDIUM risk vouchers (next 15%)", f"{n_vch_med:,}"),
        ("LOW risk vouchers", f"{n_vch_low:,}"),
        ("", None),
        ("AUDIT SAMPLE SELECTED", None),
        ("Total vouchers selected", f"{n_sel:,}"),
        ("  — HIGH risk (mandatory)", f"{n_sel_high:,}"),
        ("  — MEDIUM risk (proportional)", f"{n_sel_med:,}"),
        ("  — LOW risk (baseline)", f"{n_sel_low:,}"),
        ("Total line items in selected vouchers",
         f"{int(selected_vouchers['voucher_line_count'].sum()):,}"
         if 'voucher_line_count' in selected_vouchers.columns else "N/A"),
        ("", None),
        ("KEY FINDINGS", None),
        ("Reversal / Credit Note Transactions:", f"{n_reversals:,} (SGD {abs_rev_tot:,.2f})"),
        ("Potential FY Split Purchases:", f"{n_fy_groups:,} group(s) (SGD {fy_total:,.2f})"),
    ]

    for r_offset, (label, value) in enumerate(rows, start=3):
        label_cell = ws.cell(row=r_offset, column=1, value=label)
        is_section = value is None and label != ''
        if is_section:
            label_cell.font = Font(bold=True, color="1F3864", size=10)
            label_cell.fill = SECTION_FILL
            ws.merge_cells(f'A{r_offset}:C{r_offset}')
        elif label == '':
            pass
        else:
            label_cell.font = Font(size=10)
            val_cell = ws.cell(row=r_offset, column=2, value=value)
            val_cell.font = Font(size=10)
            val_cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'B{r_offset}:C{r_offset}')

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30

    for r in range(3, 3 + len(rows)):
        ws.row_dimensions[r].height = 15

    # T08 de-prioritisation note
    note_row = 3 + len(rows) + 1
    note_text = (
        "Note: Vendors with IDs beginning with 'T08' (typically government agencies) have been "
        "de-prioritised and will not be selected as samples unless no other vouchers are "
        "available in the lower risk tier. They remain visible in the All Vouchers Scored "
        "sheet for reference."
    )
    note_cell = ws.cell(row=note_row, column=1, value=note_text)
    note_cell.fill = PatternFill("solid", fgColor="FFC000")
    note_cell.font = Font(size=10, bold=False)
    note_cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(f'A{note_row}:C{note_row}')
    ws.row_dimensions[note_row].height = 40

    # Caveat and selection explanation blocks
    _summary_extra = [
        (
            "Scope and Limitations",
            "This tool has not been trained on confirmed fraud cases from this organisation. It "
            "identifies unusual patterns by learning the normal behaviour of the organisation's "
            "payment data, making it deployable immediately without any prior training or labelled "
            "dataset. Real-world performance depends on the nature and prevalence of anomalies "
            "present in the data. Transactions not flagged by the tool should not be interpreted "
            "as confirmation that they are free from irregularities, as sophisticated anomalies "
            "that closely mimic normal payment patterns may not be detected. Auditors should not "
            "rely on the tool to detect fraud but should exercise professional judgement in "
            "investigating unusual transactions identified.",
        ),
        (
            "Sample Selection Basis",
            "Samples are selected primarily on the basis of composite risk scores. Additional "
            "considerations are applied to ensure sample diversity: transactions with a high "
            "degree of similarity in payment description within the same vendor are deduplicated "
            "in favour of the higher-scoring voucher, and a limit of 2 samples per vendor is "
            "enforced. Vendors with IDs beginning with 'T08' (typically government agencies) are "
            "de-prioritised and will not appear in the sample unless insufficient non-T08 "
            "vouchers are available. The selected samples are intended as risk-based suggestions "
            "to guide audit focus. Auditors should exercise professional judgement in determining "
            "which payments to proceed with for further testing.",
        ),
    ]
    cur_row = note_row + 2  # leave blank gap
    for _lbl, _txt in _summary_extra:
        lc = ws.cell(row=cur_row, column=1, value=_lbl)
        lc.fill = HEADER_FILL
        lc.font = HEADER_FONT
        lc.alignment = Alignment(wrap_text=True, vertical='center')
        ws.merge_cells(f'A{cur_row}:C{cur_row}')
        ws.row_dimensions[cur_row].height = 20
        cur_row += 1
        tc = ws.cell(row=cur_row, column=1, value=_txt)
        tc.fill = ALT_FILL
        tc.font = Font(size=10)
        tc.alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{cur_row}:C{cur_row}')
        ws.row_dimensions[cur_row].height = 80
        cur_row += 2  # blank row between items


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def export_excel(df_scored, df_vouchers, selected_vouchers,
                 benford_summary, benford_stats, output_path):
    wb = Workbook()

    _sheet_selected_vouchers(wb, selected_vouchers)
    _sheet_voucher_line_detail(wb, df_scored, selected_vouchers)
    _sheet_all_vouchers(wb, df_vouchers)
    _sheet_all_lines(wb, df_scored)
    _sheet_reversals_review(wb, df_scored)
    _sheet_fy_split_purchase(wb, df_scored)
    _sheet_benford(wb, benford_summary, benford_stats)
    _sheet_summary(wb, df_scored, df_vouchers, selected_vouchers, benford_stats)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"  Excel saved: {output_path}")
